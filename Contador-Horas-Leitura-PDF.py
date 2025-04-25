import fitz  # PyMuPDF
import re
import openpyxl
from openpyxl import load_workbook
from shutil import copyfile
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# Oculta janela do tkinter
Tk().withdraw()

# Seleciona o PDF
pdf_path = askopenfilename(
    title="Selecione o arquivo PDF",
    filetypes=[("Arquivos PDF", "*.pdf")]
)
if not pdf_path:
    print("Nenhum PDF selecionado. Encerrando.")
    exit()

# Caminhos fixos do Excel
excel_original = "Horarios.xlsx"
excel_saida = "Horarios_Preenchido.xlsx"

# Verifica se existe a planilha original
if not os.path.exists(excel_original):
    print(f"Arquivo {excel_original} não encontrado.")
    exit()

# Copia a planilha modelo
copyfile(excel_original, excel_saida)

# Extrai texto do PDF
doc = fitz.open(pdf_path)
texto = ""
for page in doc:
    texto += page.get_text("text")

# Regex para capturar os dados
regex_linha = re.compile(
    r"(\d{2}/\d{2}/\d{4})\s+"         # Data de Partida
    r"(\d{2}:\d{2}:\d{2})\s+"         # Hora de Partida
    r"(\d{2}/\d{2}/\d{4})\s+"         # Data de Chegada
    r"(\d{2}:\d{2}:\d{2})\s+.*?"      # Hora de Chegada
    r"([\d]+)\.\d{2} km",             # Distância
    re.DOTALL
)

matches = regex_linha.findall(texto)

if not matches:
    print("⚠️ Nenhum dado foi encontrado no PDF.")
    exit()

# Abre a planilha e preenche os dados
wb = load_workbook(excel_saida)
ws = wb.active
linha_excel = 3  # Começa na linha 3

for match in matches:
    data_partida, hora_partida, data_chegada, hora_chegada, distancia = match

    ws[f"A{linha_excel}"] = data_partida
    ws[f"B{linha_excel}"] = hora_partida
    ws[f"D{linha_excel}"] = data_chegada
    ws[f"E{linha_excel}"] = hora_chegada
    ws[f"G{linha_excel}"] = int(distancia)

    linha_excel += 1

# Salva o Excel preenchido
wb.save(excel_saida)
print(f"\n✅ Planilha preenchida com sucesso: {excel_saida}")
